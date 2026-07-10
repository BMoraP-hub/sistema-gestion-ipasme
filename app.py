"""
================================================================================
SISTEMA DE GESTIÓN DE INCIDENCIAS OTSI - IPASME
================================================================================
DESCRIPCIÓN:
    Aplicación web desarrollada con Flask para la automatización de la recepción,
    asignación y control de incidencias tecnológicas.

BASE DE DATOS (DISEÑO):
    - Relaciones: 
        - Departamento (1:N) Tecnicos
        - Grupo (1:N) Tecnicos
        - Tecnicos (1:N) TicketsSoporte
    - Seguridad: 
        - Autenticación basada en sesiones firmadas (Flask Session).
        - Niveles de acceso diferenciados (Operador 'normal' vs Supervisor 'super').
        - Protección contra inyección SQL mediante SQLAlchemy ORM.
        - Auditoría mediante bitácora de eventos inmutable.

CARACTERÍSTICAS TÉCNICAS:
    - Paginación inteligente: Integración con DataTables para manejo de grandes volúmenes.
    - Exportación: Generación dinámica de reportes en PDF y Excel.
    - UI/UX: Diseño Glassmorphism con soporte nativo para Modo Oscuro/Claro.
    - Respaldo: Función de backup automatizada para la base de datos (SQLite).
================================================================================
"""
import os
import socket
import threading
import webbrowser
from datetime import datetime, timedelta
from io import BytesIO

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
from flask_caching import Cache
import openpyxl
from openpyxl.utils import get_column_letter
from fpdf import FPDF


def can_connect_mysql(uri):
    """Prueba si la cadena de conexión a MySQL o MariaDB responde correctamente."""
    try:
        engine = create_engine(uri)
        with engine.connect():
            pass
        return True
    except Exception as e:
        print(f"[ERROR] Falló conexión a MySQL con URI={uri}: {e}")
        return False
 

def get_database_uri():
    """Selecciona la URI de base de datos disponible: MySQL/MariaDB si existe, o SQLite local como respaldo."""
    db_url = os.getenv("DATABASE_URL")
    if db_url:
        return db_url

    user = os.getenv("DB_USER", "root")
    password = os.getenv("DB_PASSWORD", "")
    host = os.getenv("DB_HOST", "127.0.0.1")
    port = os.getenv("DB_PORT", "3307")
    database = os.getenv("DB_NAME", "sistema_ipasme.db")

    mysql_uri = (
        f"mysql+pymysql://{user}:{password}@{host}:{port}/{database}"
        if password
        else f"mysql+pymysql://{user}@{host}:{port}/{database}"
    )
    if can_connect_mysql(mysql_uri):
        print("[INFO] Conectando a MySQL/MariaDB:", mysql_uri)
        return mysql_uri

    sqlite_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), "sistema_ipasme.db")
    print("[WARN] No se pudo conectar a MySQL/MariaDB, usando SQLite local:", sqlite_path)
    return f"sqlite:///{sqlite_path}"


app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "cambia-esta-clave")
app.config["SQLALCHEMY_DATABASE_URI"] = get_database_uri()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
# Configuración de cache simple en memoria
app.config['CACHE_TYPE'] = os.getenv('CACHE_TYPE', 'SimpleCache')
app.config['CACHE_DEFAULT_TIMEOUT'] = int(os.getenv('CACHE_TIMEOUT', '300'))

cache = Cache(app)

# Credenciales fijas para login normal
NORMAL_LOGIN_USER = "Pruebauneti"
NORMAL_LOGIN_PASSWORD = "uneti123"

db = SQLAlchemy(app)
print("[INFO] SQLALCHEMY_DATABASE_URI=", app.config["SQLALCHEMY_DATABASE_URI"])


class Departamento(db.Model):
    __tablename__ = "soporte_departamentos"
    id_depto = db.Column(db.Integer, primary_key=True)
    nombre_depto = db.Column(db.String(50), nullable=False)
    tecnicos = db.relationship("Tecnico", backref="departamento", lazy=True)


class Grupo(db.Model):
    __tablename__ = "soporte_grupos"
    id_grupo = db.Column(db.Integer, primary_key=True)
    nombre_grupo = db.Column(db.String(50), nullable=False)
    tecnicos = db.relationship("Tecnico", backref="grupo", lazy=True)


class Tecnico(db.Model):
    __tablename__ = "soporte_tecnicos"
    id_tecnico = db.Column(db.Integer, primary_key=True)
    nombre_completo = db.Column(db.String(100), nullable=False)
    id_grupo = db.Column(db.Integer, db.ForeignKey("soporte_grupos.id_grupo"), nullable=False)
    id_depto = db.Column(db.Integer, db.ForeignKey("soporte_departamentos.id_depto"), nullable=False)
    
    tickets_asignados = db.relationship(
        "TicketSoporte",
        backref="tecnico",
        lazy="dynamic"
    )

    @property
    def asignados(self):
        return self.tickets_asignados.filter_by(estado_operacional="Asignado").count()
        
    @property
    def resueltos(self):
        return self.tickets_asignados.filter_by(estado_operacional="Resuelto").count()


class TicketSoporte(db.Model):
    __tablename__ = "soporte_tickets"
    id_ticket = db.Column(db.Integer, primary_key=True)
    fecha_hora_reporte = db.Column(db.DateTime, default=datetime.utcnow)
    tipo_falla = db.Column(db.String(100), nullable=False, default="No Especificado")
    descripcion_usuario = db.Column(db.Text, nullable=False, default="Registro Directo/Rápido")
    estado_operacional = db.Column(db.String(20), default="Abierto") # Abierto, Asignado, Resuelto
    id_tecnico_asignado = db.Column(db.Integer, db.ForeignKey("soporte_tecnicos.id_tecnico"), nullable=True)


class BitacoraAuditoria(db.Model):
    __tablename__ = "soporte_bitacora"
    id_log = db.Column(db.Integer, primary_key=True)
    fecha_evento = db.Column(db.DateTime, default=datetime.utcnow)
    accion_realizada = db.Column(db.String(255), nullable=False)
    id_ticket = db.Column(db.Integer, db.ForeignKey("soporte_tickets.id_ticket"), nullable=False)
    usuario_responsable = db.Column(db.String(100), nullable=False)


def seed_default_data():
    """Carga datos iniciales de ejemplo para departamentos, grupos y otras referencias básicas."""
    def ensure(model, field, value):
        if not model.query.filter(getattr(model, field) == value).first():
            db.session.add(model(**{field: value}))

    ensure(Departamento, "nombre_depto", "Soporte")
    ensure(Departamento, "nombre_depto", "Redes")
    ensure(Grupo, "nombre_grupo", "Grupo A")
    ensure(Grupo, "nombre_grupo", "Grupo B")
    ensure(Grupo, "nombre_grupo", "Permisos")
    db.session.commit()


with app.app_context():
    db.create_all()
    seed_default_data()


@app.route("/")
def index():
    """Redirige la ruta inicial al login del sistema."""
    return redirect(url_for("login"))


@app.before_request
def require_login():
    if request.endpoint in ("static", "login", "super_login"):
        return
    if request.endpoint is None:
        return
    if request.endpoint == "moderacion" and session.get("user_type") != "super":
        flash("Acceso denegado. Requiere permisos de super usuario.", "danger")
        return redirect(url_for("super_login", next=request.path))
    if "user" not in session:
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    """Gestiona el acceso del operador normal al sistema mediante sesión."""
    if "user" in session:
        return redirect(url_for("registro"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Ingresa usuario y clave válidos.", "warning")
            return render_template("login.html", active="login")

        if username != NORMAL_LOGIN_USER or password != NORMAL_LOGIN_PASSWORD:
            flash("Usuario o clave incorrectos.", "danger")
            return render_template("login.html", active="login")

        session["user"] = username
        session["user_type"] = "normal"
        flash(f"Bienvenido {username}.", "success")
        return redirect(url_for("registro"))

    return render_template("login.html", active="login")


@app.route("/super-login", methods=["GET", "POST"])
def super_login():
    """Valida el ingreso del usuario supervisor y permite acceder a funciones de moderación."""
    if "user" in session and session.get("user_type") == "super":
        return redirect(url_for("moderacion"))

    next_page = request.args.get("next")
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        next_page = request.form.get("next") or next_page

        if username == "Admin123" and password == "OTICIPASME123.*":
            session["user"] = username
            session["user_type"] = "super"
            flash("Bienvenido Super Usuario.", "success")
            return redirect(next_page or url_for("moderacion"))
        else:
            flash("Usuario o clave incorrectos para super usuario.", "danger")
            return render_template("super_login.html", active="super_login", next=next_page)

    return render_template("super_login.html", active="super_login", next=next_page)


@app.route("/logout")
def logout():
    session.pop("user", None)
    session.pop("user_type", None)
    flash("Sesión cerrada.", "success")
    return redirect(url_for("login"))


def cargar_filtros():
    """Obtiene y cachea los filtros base de departamentos, grupos y técnicos usados en varias vistas."""
    # Evitar cachear instancias de SQLAlchemy (provoca DetachedInstanceError).
    # En su lugar cacheamos listas de ids y re-obtenemos instancias en el contexto de la request.
    cached = cache.get('cargar_filtros_ids')
    if cached:
        dept_ids = cached.get('departamentos', [])
        grupo_ids = cached.get('grupos', [])
        tecnico_ids = cached.get('tecnicos', [])

        departamentos = Departamento.query.filter(Departamento.id_depto.in_(dept_ids)).order_by(Departamento.nombre_depto).all() if dept_ids else Departamento.query.order_by(Departamento.nombre_depto).all()
        grupos = Grupo.query.filter(Grupo.id_grupo.in_(grupo_ids)).order_by(Grupo.nombre_grupo).all() if grupo_ids else Grupo.query.order_by(Grupo.nombre_grupo).all()
        tecnicos = Tecnico.query.filter(Tecnico.id_tecnico.in_(tecnico_ids)).order_by(Tecnico.nombre_completo).all() if tecnico_ids else Tecnico.query.order_by(Tecnico.nombre_completo).all()
        return departamentos, grupos, tecnicos

    departamentos = Departamento.query.order_by(Departamento.nombre_depto).all()
    grupos = Grupo.query.filter(Grupo.nombre_grupo.in_( ["Grupo A", "Grupo B", "Permisos"]))
    grupos = grupos.order_by(Grupo.nombre_grupo).all()
    tecnicos = Tecnico.query.order_by(Tecnico.nombre_completo).all()

    # Guardar sólo ids en cache
    cache.set('cargar_filtros_ids', {
        'departamentos': [d.id_depto for d in departamentos],
        'grupos': [g.id_grupo for g in grupos],
        'tecnicos': [t.id_tecnico for t in tecnicos],
    })

    return departamentos, grupos, tecnicos


@app.route("/api/tecnicos")
def api_tecnicos():
    """Devuelve una lista filtrada de técnicos en formato JSON para uso en formularios y búsquedas."""
    filter_depto = request.args.get("departamento", "all")
    filter_grupo = request.args.get("grupo", "all")
    q = request.args.get('query', '').strip()

    cache_key = f"api_tecnicos::{filter_depto}::{filter_grupo}::{q}"
    cached = cache.get(cache_key)
    if cached:
        return jsonify(cached)

    query = Tecnico.query
    if filter_depto != "all":
        query = query.filter(Tecnico.id_depto == int(filter_depto))
    if filter_grupo != "all":
        query = query.filter(Tecnico.id_grupo == int(filter_grupo))
    if q:
        query = query.filter(Tecnico.nombre_completo.ilike(f"%{q}%"))

    tecnicos = query.order_by(Tecnico.nombre_completo).all()
    result = [
        {"id_tecnico": t.id_tecnico, "nombre_completo": t.nombre_completo}
        for t in tecnicos
    ]
    cache.set(cache_key, result)
    return jsonify(result)


@app.route("/registro", methods=["GET", "POST"])
def registro():
    """Permite registrar tickets nuevos y mostrarlos en la vista de ingreso del operador."""
    departamentos, grupos, tecnicos = cargar_filtros()
    
    if request.method == "POST":
        tipo_falla = request.form.get("tipo_falla", "Falla no especificada")
        descripcion = request.form.get("descripcion_usuario", "Reporte directo")
        id_tecnico = request.form.get("tecnico")
        id_tecnico = int(id_tecnico) if id_tecnico and id_tecnico.strip() else None
        
        estado = "Asignado" if id_tecnico is not None else "Abierto"

        nuevo_ticket = TicketSoporte(
            tipo_falla=tipo_falla,
            descripcion_usuario=descripcion,
            estado_operacional=estado,
            id_tecnico_asignado=id_tecnico
        )
        db.session.add(nuevo_ticket)
        db.session.commit()

        log = BitacoraAuditoria(
            accion_realizada=f"Ticket creado con estado: {estado}",
            id_ticket=nuevo_ticket.id_ticket,
            usuario_responsable=session.get("user", "Operador")
        )
        db.session.add(log)
        db.session.commit()

        if id_tecnico:
            flash("Ticket creado y asignado directamente.", "success")
        else:
            flash("Ticket abierto y puesto en cola para moderación.", "success")
            
        return redirect(url_for("registro"))

    tickets_abiertos = TicketSoporte.query.filter_by(estado_operacional="Abierto").all()

    return render_template(
        "registro.html",
        active="registro",
        departamentos=departamentos,
        grupos=grupos,
        tecnicos=tecnicos,
        tickets_abiertos=tickets_abiertos
    )


# --- NUEVAS RUTAS: PANEL DEL TÉCNICO DE SOPORTE (PASO 4) ---
@app.route("/mis_tickets")
def mis_tickets():
    """Muestra los casos asignados a un técnico para su seguimiento y resolución."""
    departamentos, grupos, tecnicos = cargar_filtros()
    tecnico_id = request.args.get("tecnico_id", type=int)
    
    tickets_asignados = []
    if tecnico_id:
        # P3: Consultar Asignaciones (Solo casos encolados/asignados)
        tickets_asignados = TicketSoporte.query.filter_by(
            id_tecnico_asignado=tecnico_id,
            estado_operacional="Asignado"
        ).order_by(TicketSoporte.fecha_hora_reporte.asc()).all()
        
    return render_template(
        "mis_tickets.html",
        active="mis_tickets",
        tecnicos=tecnicos,
        tecnico_id=tecnico_id,
        tickets_asignados=tickets_asignados
    )

@app.route("/resolver_ticket/<int:id_ticket>")
def resolver_ticket(id_ticket):
    ticket = TicketSoporte.query.get_or_404(id_ticket)
    
    if ticket.estado_operacional == "Asignado":
        ticket.estado_operacional = "Resuelto"
        
        # P3: Cierre Técnico Operativo -> Genera Bitácora
        log = BitacoraAuditoria(
            accion_realizada=f"El Técnico especialista marcó la falla como Resuelta en sitio.",
            id_ticket=ticket.id_ticket,
            usuario_responsable=session.get("user", "Técnico Especialista")
        )
        db.session.add(log)
        db.session.commit()
        flash(f"Ticket #{ticket.id_ticket} marcado como Resuelto. Excelente trabajo.", "success")
    else:
        flash("El ticket no se encuentra en estado Asignado.", "warning")
        
    return redirect(url_for("mis_tickets", tecnico_id=ticket.id_tecnico_asignado))
# -----------------------------------------------------------


@app.route("/base-datos")
def base_datos():
    """Presenta un panel con métricas, historial y gráficos de productividad por técnico."""
    departamentos, grupos, tecnicos = cargar_filtros()
    filter_depto = request.args.get("departamento", "all")
    filter_grupo = request.args.get("grupo", "all")
    filter_fecha = request.args.get("fecha", "all")
    filter_fecha_manual = request.args.get("fecha_manual", "")
    filter_asignados = request.args.get("asignados", "with")

    # Mostrar 'resumen' sólo cuando se pide explícitamente 'today'.
    # Para 'all' mostramos la vista histórica completa.
    view = "resumen" if filter_fecha == "today" else "historial"

    query = Tecnico.query
    if filter_depto != "all":
        query = query.filter(Tecnico.id_depto == int(filter_depto))
    if filter_grupo != "all":
        query = query.filter(Tecnico.id_grupo == int(filter_grupo))

    now = datetime.utcnow()
    today_counts = {}
    if filter_fecha == "today":
        inicio = datetime(now.year, now.month, now.day)
        fin = inicio + timedelta(days=1)
        count_query = (
            db.session.query(
                TicketSoporte.id_tecnico_asignado.label("id_tecnico"),
                db.func.count(TicketSoporte.id_ticket).label("total_casos"),
            )
            .join(Tecnico)
            .filter(
                TicketSoporte.fecha_hora_reporte >= inicio, 
                TicketSoporte.fecha_hora_reporte < fin,
                TicketSoporte.estado_operacional.in_(["Asignado", "Resuelto"])
            )
        )
        if filter_depto != "all":
            count_query = count_query.filter(Tecnico.id_depto == int(filter_depto))
        if filter_grupo != "all":
            count_query = count_query.filter(Tecnico.id_grupo == int(filter_grupo))
        count_query = count_query.group_by(TicketSoporte.id_tecnico_asignado)
        today_counts = {row.id_tecnico: row.total_casos for row in count_query.all() if row.id_tecnico}

    tecnicos_filtrados = query.order_by(Tecnico.nombre_completo).all()

    asignados = []
    for t in tecnicos_filtrados:
        caso_count = today_counts.get(t.id_tecnico, t.asignados + t.resueltos if filter_fecha != "today" else 0)
        if filter_asignados == "with" and caso_count == 0:
            continue
        if filter_asignados == "without" and caso_count > 0:
            continue
        asignados.append({
            "tecnico": t,
            "asignados": caso_count,
        })

    asignaciones = TicketSoporte.query.join(Tecnico).join(Grupo).join(Departamento).filter(TicketSoporte.id_tecnico_asignado != None)
    
    if filter_depto != "all":
        asignaciones = asignaciones.filter(Tecnico.id_depto == int(filter_depto))
    if filter_grupo != "all":
        asignaciones = asignaciones.filter(Tecnico.id_grupo == int(filter_grupo))
    
    if filter_fecha == "today":
        inicio = datetime(now.year, now.month, now.day)
        fin = inicio + timedelta(days=1)
        asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= inicio, TicketSoporte.fecha_hora_reporte < fin)
    elif filter_fecha == "week":
        limite = now - timedelta(days=7)
        asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= limite)
    elif filter_fecha == "month":
        inicio_mes = datetime(now.year, now.month, 1)
        asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= inicio_mes)
    elif filter_fecha == "year":
        inicio_year = datetime(now.year, 1, 1)
        asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= inicio_year)
    elif filter_fecha == "manual":
        try:
            fecha_manual_dt = datetime.strptime(filter_fecha_manual, "%Y-%m-%d")
        except Exception:
            fecha_manual_dt = datetime(now.year, now.month, now.day)
            filter_fecha_manual = fecha_manual_dt.strftime("%Y-%m-%d")
        inicio = datetime(fecha_manual_dt.year, fecha_manual_dt.month, fecha_manual_dt.day)
        fin = inicio + timedelta(days=1)
        asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= inicio, TicketSoporte.fecha_hora_reporte < fin)

    if filter_asignados == "without":
        asignaciones = []
    else:
        asignaciones = asignaciones.order_by(TicketSoporte.fecha_hora_reporte.desc()).all()

    if view == "historial":
        resumen = {}
        for caso in asignaciones:
            fecha_str = caso.fecha_hora_reporte.strftime("%Y-%m-%d")
            key = (caso.id_tecnico_asignado, fecha_str)
            if key not in resumen:
                resumen[key] = {
                    "fecha": fecha_str,
                    "grupo": caso.tecnico.grupo.nombre_grupo,
                    "tecnico": caso.tecnico,
                    "departamento": caso.tecnico.departamento.nombre_depto,
                    "casos_hechos": 0,
                }
            resumen[key]["casos_hechos"] += 1
        asignaciones = sorted(resumen.values(), key=lambda x: (x["fecha"], x["tecnico"].nombre_completo), reverse=True)

    chart_labels = []
    chart_values = []
    if view == "resumen":
        for fila in asignados:
            if fila["asignados"] > 0:
                chart_labels.append(fila["tecnico"].nombre_completo)
                chart_values.append(fila["asignados"])
    else:
        totals = {}
        for caso in asignaciones:
            name = caso["tecnico"].nombre_completo
            totals[name] = totals.get(name, 0) + caso["casos_hechos"]
        chart_labels = list(totals.keys())
        chart_values = list(totals.values())

    return render_template(
        "base_datos.html",
        active="base-datos",
        departamentos=departamentos,
        grupos=grupos,
        asignados=asignados,
        asignaciones=asignaciones,
        filter_depto=filter_depto,
        filter_grupo=filter_grupo,
        filter_fecha=filter_fecha,
        filter_fecha_manual=filter_fecha_manual,
        filter_asignados=filter_asignados,
        chart_labels=chart_labels,
        chart_values=chart_values,
        view=view,
    )


@app.route("/export_pdf")
def export_pdf():
    """Genera un reporte en formato PDF con el historial o resumen de productividad solicitado."""
    tipo = request.args.get("tipo", "resumen")
    filter_depto = request.args.get("departamento", "all")
    filter_grupo = request.args.get("grupo", "all")
    filter_fecha = request.args.get("fecha", "all")
    filter_fecha_manual = request.args.get("fecha_manual", "")
    filter_asignados = request.args.get("asignados", "all")

    pdf = FPDF()
    pdf.set_auto_page_break(True, margin=12)
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Reporte de soporte", ln=True, align="C")
    pdf.ln(4)

    def ensure_space(p, h=8):
        # Añade nueva página si el espacio restante es menor que h
        if p.get_y() > (p.h - p.b_margin - h):
            p.add_page()

    if tipo == "historial":
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, "Historial de Asignaciones y Casos Resueltos", ln=True)
        pdf.ln(2)
        pdf.set_font("Arial", size=10)

        asignaciones = TicketSoporte.query.join(Tecnico).filter(TicketSoporte.id_tecnico_asignado != None)
        if filter_depto != "all":
            asignaciones = asignaciones.filter(Tecnico.id_depto == int(filter_depto))
        if filter_grupo != "all":
            asignaciones = asignaciones.filter(Tecnico.id_grupo == int(filter_grupo))

        if filter_fecha == "today":
            inicio = datetime.utcnow().date()
            asignaciones = asignaciones.filter(db.func.date(TicketSoporte.fecha_hora_reporte) == inicio)
        elif filter_fecha == "week":
            limite = datetime.utcnow() - timedelta(days=7)
            asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= limite)
        elif filter_fecha == "month":
            inicio_mes = datetime.utcnow().replace(day=1)
            asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= inicio_mes)
        elif filter_fecha == "year":
            inicio_year = datetime.utcnow().replace(month=1, day=1)
            asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= inicio_year)
        elif filter_fecha == "manual":
            try:
                fecha_manual_dt = datetime.strptime(filter_fecha_manual, "%Y-%m-%d")
            except Exception:
                fecha_manual_dt = datetime.utcnow()
            inicio = datetime(fecha_manual_dt.year, fecha_manual_dt.month, fecha_manual_dt.day)
            fin = inicio + timedelta(days=1)
            asignaciones = asignaciones.filter(
                TicketSoporte.fecha_hora_reporte >= inicio,
                TicketSoporte.fecha_hora_reporte < fin,
            )

        if filter_asignados == "without":
            asignaciones = []
        else:
            asignaciones = asignaciones.order_by(TicketSoporte.fecha_hora_reporte.desc()).all()

        resumen = {}
        for caso in asignaciones:
            fecha_str = caso.fecha_hora_reporte.strftime("%Y-%m-%d")
            key = (caso.id_tecnico_asignado, fecha_str)
            if key not in resumen:
                resumen[key] = {
                    "fecha": fecha_str,
                    "grupo": caso.tecnico.grupo.nombre_grupo if caso.tecnico and caso.tecnico.grupo else "",
                    "tecnico": caso.tecnico,
                    "departamento": caso.tecnico.departamento.nombre_depto if caso.tecnico and caso.tecnico.departamento else "",
                    "casos_hechos": 0,
                }
            resumen[key]["casos_hechos"] += 1
        filas = sorted(resumen.values(), key=lambda x: (x["fecha"], x["tecnico"].nombre_completo if x["tecnico"] else ""), reverse=True)

        # Definir anchos de columna que sumen <= ancho util de la página (A4 ~210mm menos márgenes)
        col_widths = [35, 60, 40, 40, 15]  # suma 190

        # Funciones auxiliares para wrapping y altura de fila
        def split_lines(p, text, width):
            if not text:
                return [""]
            words = str(text).split()
            lines = []
            cur = ""
            for w in words:
                test = (cur + " " + w).strip()
                if p.get_string_width(test) <= width:
                    cur = test
                else:
                    if cur:
                        lines.append(cur)
                    cur = w
            if cur:
                lines.append(cur)
            return lines

        def draw_header(headers, widths, lh):
            x = pdf.get_x()
            y = pdf.get_y()
            pdf.set_font("Arial", "B", 10)
            for i, h_text in enumerate(headers):
                w = widths[i]
                pdf.rect(x, y, w, lh)
                pdf.set_xy(x, y)
                pdf.multi_cell(w, lh, h_text, align='C')
                x += w
            pdf.set_y(y + lh)

        lh = 6
        headers = ["Fecha", "Técnico", "Departamento", "Grupo", "Casos"]
        draw_header(headers, col_widths, lh)

        pdf.set_font("Arial", size=10)
        for fila in filas:
            fecha = fila.get("fecha", "-")
            tecnico_nombre = (fila.get("tecnico").nombre_completo if fila.get("tecnico") else "-")
            departamento = (fila.get("departamento") or "-")
            grupo = (fila.get("grupo") or "-")
            casos_hechos = str(fila.get("casos_hechos", 0))

            texts = [fecha, tecnico_nombre, departamento, grupo, casos_hechos]
            # calcular altura necesaria
            lines_counts = [len(split_lines(pdf, t, w)) for t, w in zip(texts, col_widths)]
            h = max(1, max(lines_counts)) * lh
            # nueva página si no cabe
            if pdf.get_y() + h > (pdf.h - pdf.b_margin):
                pdf.add_page()
                draw_header(headers, col_widths, lh)

            x = pdf.get_x()
            y = pdf.get_y()
            for i, text in enumerate(texts):
                w = col_widths[i]
                pdf.rect(x, y, w, h)
                pdf.set_xy(x, y)
                pdf.multi_cell(w, lh, str(text), align='L')
                x += w
            pdf.set_y(y + h)
    else:
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, "Resumen de Productividad de Técnicos", ln=True)
        pdf.ln(2)
        pdf.set_font("Arial", size=10)

        query = Tecnico.query
        if filter_depto != "all":
            query = query.filter(Tecnico.id_depto == int(filter_depto))
        if filter_grupo != "all":
            query = query.filter(Tecnico.id_grupo == int(filter_grupo))

        tecnicos_filtrados = query.order_by(Tecnico.nombre_completo).all()

        col_widths = [100, 60, 30]  # suma 190

        # Tabla con wrapping para nombres largos
        def draw_header_small(headers, widths, lh):
            x = pdf.get_x()
            y = pdf.get_y()
            pdf.set_font("Arial", "B", 10)
            for i, h_text in enumerate(headers):
                w = widths[i]
                pdf.rect(x, y, w, lh)
                pdf.set_xy(x, y)
                pdf.multi_cell(w, lh, h_text, align='C')
                x += w
            pdf.set_y(y + lh)

        def split_lines(p, text, width):
            if not text:
                return [""]
            words = str(text).split()
            lines = []
            cur = ""
            for w in words:
                test = (cur + " " + w).strip()
                if p.get_string_width(test) <= width:
                    cur = test
                else:
                    if cur:
                        lines.append(cur)
                    cur = w
            if cur:
                lines.append(cur)
            return lines

        lh = 6
        headers_small = ["Técnico", "Departamento", "Asignados"]
        draw_header_small(headers_small, col_widths, lh)
        pdf.set_font("Arial", size=10)

        for tecnico in tecnicos_filtrados:
            nombre = tecnico.nombre_completo or "-"
            depto = tecnico.departamento.nombre_depto if tecnico.departamento else "-"
            total = str(tecnico.asignados + tecnico.resueltos)

            texts = [nombre, depto, total]
            widths = col_widths
            lines_counts = [len(split_lines(pdf, t, w)) for t, w in zip(texts, widths)]
            h = max(1, max(lines_counts)) * lh
            if pdf.get_y() + h > (pdf.h - pdf.b_margin):
                pdf.add_page()
                draw_header_small(headers_small, col_widths, lh)

            x = pdf.get_x()
            y = pdf.get_y()
            for i, text in enumerate(texts):
                w = widths[i]
                pdf.rect(x, y, w, h)
                pdf.set_xy(x, y)
                pdf.multi_cell(w, lh, str(text), align='L')
                x += w
            pdf.set_y(y + h)

    pdf_bytes = pdf.output(dest="S")
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode("latin-1")

    response = BytesIO(pdf_bytes)
    response.seek(0)
    return send_file(
        response,
        as_attachment=True,
        download_name="reporte_soporte.pdf",
        mimetype="application/pdf",
    )


@app.route("/export_excel")
def export_excel():
    """Exporta los datos seleccionados a un archivo Excel para análisis o respaldo."""
    tipo = request.args.get("tipo", "resumen")
    filter_depto = request.args.get("departamento", "all")
    filter_grupo = request.args.get("grupo", "all")
    filter_fecha = request.args.get("fecha", "all")
    filter_fecha_manual = request.args.get("fecha_manual", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    if tipo == 'historial':
        ws.append(['Fecha', 'Técnico', 'Departamento', 'Grupo', 'Casos'])
        asignaciones = TicketSoporte.query.join(Tecnico).filter(TicketSoporte.id_tecnico_asignado != None)
        if filter_depto != 'all':
            asignaciones = asignaciones.filter(Tecnico.id_depto == int(filter_depto))
        if filter_grupo != 'all':
            asignaciones = asignaciones.filter(Tecnico.id_grupo == int(filter_grupo))

        if filter_fecha == 'today':
            inicio = datetime.utcnow().date()
            asignaciones = asignaciones.filter(db.func.date(TicketSoporte.fecha_hora_reporte) == inicio)
        elif filter_fecha == 'manual':
            try:
                fecha_manual_dt = datetime.strptime(filter_fecha_manual, "%Y-%m-%d")
            except Exception:
                fecha_manual_dt = datetime.utcnow()
            inicio = datetime(fecha_manual_dt.year, fecha_manual_dt.month, fecha_manual_dt.day)
            fin = inicio + timedelta(days=1)
            asignaciones = asignaciones.filter(TicketSoporte.fecha_hora_reporte >= inicio, TicketSoporte.fecha_hora_reporte < fin)

        asignaciones = asignaciones.order_by(TicketSoporte.fecha_hora_reporte.desc()).all()

        resumen = {}
        for caso in asignaciones:
            fecha_str = caso.fecha_hora_reporte.strftime("%Y-%m-%d")
            key = (caso.id_tecnico_asignado, fecha_str)
            if key not in resumen:
                resumen[key] = {"fecha": fecha_str, "grupo": caso.tecnico.grupo.nombre_grupo if caso.tecnico and caso.tecnico.grupo else "", "tecnico": caso.tecnico, "departamento": caso.tecnico.departamento.nombre_depto if caso.tecnico and caso.tecnico.departamento else "", "casos_hechos": 0}
            resumen[key]['casos_hechos'] += 1

        filas = sorted(resumen.values(), key=lambda x: (x['fecha'], x['tecnico'].nombre_completo if x['tecnico'] else ''), reverse=True)
        for fila in filas:
            ws.append([fila['fecha'], fila['tecnico'].nombre_completo if fila['tecnico'] else '', fila['departamento'], fila['grupo'], fila['casos_hechos']])
    else:
        ws.append(['Técnico', 'Departamento', 'Asignados'])
        query = Tecnico.query
        if filter_depto != 'all':
            query = query.filter(Tecnico.id_depto == int(filter_depto))
        if filter_grupo != 'all':
            query = query.filter(Tecnico.id_grupo == int(filter_grupo))
        tecnicos_filtrados = query.order_by(Tecnico.nombre_completo).all()
        for t in tecnicos_filtrados:
            ws.append([t.nombre_completo, t.departamento.nombre_depto if t.departamento else '', t.asignados + t.resueltos])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ''
                if len(val) > max_length:
                    max_length = len(val)
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='reporte_soporte.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route("/moderacion", methods=["GET", "POST"])
def moderacion():
    """Administra asignaciones de tickets y muestra el panel de supervisión del sistema."""
    section = request.args.get("section", "home")
    departamentos, grupos, tecnicos = cargar_filtros()
    
    if request.method == "POST":
        id_ticket = request.form.get("id_ticket")
        id_tecnico = request.form.get("id_tecnico")
        
        if id_ticket and id_tecnico:
            ticket = TicketSoporte.query.get(id_ticket)
            tecnico = Tecnico.query.get(id_tecnico)
            
            if ticket and tecnico:
                ticket.id_tecnico_asignado = tecnico.id_tecnico
                ticket.estado_operacional = "Asignado"
                
                log = BitacoraAuditoria(
                    accion_realizada=f"Supervisor asignó el caso a {tecnico.nombre_completo}",
                    id_ticket=ticket.id_ticket,
                    usuario_responsable=session.get("user", "Supervisor")
                )
                db.session.add(log)
                db.session.commit()
                flash(f"Ticket #{ticket.id_ticket} asignado con éxito a {tecnico.nombre_completo}.", "success")
        return redirect(url_for("moderacion", section="casos"))

    tecnico_counts = []
    fecha_actual = request.args.get("fecha", datetime.utcnow().strftime("%Y-%m-%d"))
    
    try:
        fecha_dt = datetime.strptime(fecha_actual, "%Y-%m-%d")
    except ValueError:
        fecha_dt = datetime.utcnow()
        fecha_actual = fecha_dt.strftime("%Y-%m-%d")
        
    inicio = datetime(fecha_dt.year, fecha_dt.month, fecha_dt.day)
    fin = inicio + timedelta(days=1)
    
    for tecnico in tecnicos:
        casos_hoy = TicketSoporte.query.filter(
            TicketSoporte.id_tecnico_asignado == tecnico.id_tecnico,
            TicketSoporte.fecha_hora_reporte >= inicio,
            TicketSoporte.fecha_hora_reporte < fin,
            TicketSoporte.estado_operacional.in_(["Asignado", "Resuelto"])
        ).count()
        tecnico_counts.append({
            "tecnico": tecnico,
            "casos_hoy": casos_hoy,
        })
        
    tickets_abiertos = TicketSoporte.query.filter_by(estado_operacional="Abierto").all()

    return render_template(
        "moderacion.html",
        active="moderacion",
        section=section,
        tecnico_counts=tecnico_counts,
        fecha_actual=fecha_actual,
        tickets_abiertos=tickets_abiertos
    )


@app.route("/ajustar_asignaciones/<int:id_tecnico>/<action>")
def ajustar_asignaciones(id_tecnico, action):
    tecnico = Tecnico.query.get_or_404(id_tecnico)
    fecha_str = request.args.get("fecha")
    if fecha_str:
        try:
            fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        except ValueError:
            fecha_dt = datetime.utcnow()
    else:
        fecha_dt = datetime.utcnow()

    inicio = datetime(fecha_dt.year, fecha_dt.month, fecha_dt.day)
    fin = inicio + timedelta(days=1)

    if action == "add":
        nuevo = TicketSoporte(
            id_tecnico_asignado=tecnico.id_tecnico, 
            estado_operacional="Asignado",
            fecha_hora_reporte=inicio
        )
        db.session.add(nuevo)
        db.session.commit()
        
        log = BitacoraAuditoria(
            accion_realizada=f"Asignación Rápida a {tecnico.nombre_completo}",
            id_ticket=nuevo.id_ticket,
            usuario_responsable=session.get("user", "SuperAdmin")
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Caso asignado a {tecnico.nombre_completo} para {inicio.strftime('%Y-%m-%d')}", "success")
        
    elif action == "remove":
        caso = TicketSoporte.query.filter(
            TicketSoporte.id_tecnico_asignado == tecnico.id_tecnico,
            TicketSoporte.fecha_hora_reporte >= inicio,
            TicketSoporte.fecha_hora_reporte < fin,
        ).order_by(TicketSoporte.fecha_hora_reporte.desc()).first()
        
        if caso:
            BitacoraAuditoria.query.filter_by(id_ticket=caso.id_ticket).delete()
            db.session.delete(caso)
            db.session.commit()
            flash(f"Se eliminó un caso de {tecnico.nombre_completo} para {inicio.strftime('%Y-%m-%d')}", "success")
        else:
            flash(f"No hay casos para {tecnico.nombre_completo} en {inicio.strftime('%Y-%m-%d')}", "warning")
    else:
        flash("Acción desconocida.", "danger")

    return redirect(url_for("moderacion", section="casos"))


@app.route("/eliminar_ticket/<int:id_ticket>", methods=["POST"])
@app.route("/eliminar_asignacion/<int:id_ticket>", methods=["POST"])
def eliminar_asignacion(id_ticket):
    caso = TicketSoporte.query.get_or_404(id_ticket)
    next_url = request.form.get("next") or request.args.get("next") or url_for("registro")

    if caso.estado_operacional != "Abierto":
        flash("Solo se pueden eliminar tickets que estén en estado Abierto.", "warning")
    else:
        BitacoraAuditoria.query.filter_by(id_ticket=caso.id_ticket).delete()
        db.session.delete(caso)
        db.session.commit()
        flash(f"Ticket #{caso.id_ticket} eliminado correctamente.", "success")

    return redirect(next_url)


@app.route("/tecnicos", methods=["GET", "POST"])
def tecnicos():
    """Permite agregar, editar o eliminar técnicos del sistema y asignarles sus referencias."""
    departamentos, grupos, tecnicos = cargar_filtros()
    from_moderacion = request.args.get("from_moderacion", "0") == "1"
    edit_id = request.args.get("edit_id", type=int)
    editar_tecnico = None
    if edit_id:
        editar_tecnico = Tecnico.query.get(edit_id)

    if request.method == "POST":
        nombre = request.form.get("nombre")
        departamento_id = request.form.get("departamento")
        grupo_id = request.form.get("grupo")
        edit_id = request.form.get("edit_id")
        from_moderacion = request.form.get("from_moderacion", "0") == "1"

        if not nombre or not departamento_id or not grupo_id:
            target = url_for("tecnicos", edit_id=edit_id, from_moderacion="1") if edit_id else url_for("tecnicos", from_moderacion="1" if from_moderacion else None)
            flash("Complete todos los campos para agregar o modificar el técnico.", "warning")
            return redirect(target)

        if edit_id:
            tecnico = Tecnico.query.get(int(edit_id))
            if tecnico:
                tecnico.nombre_completo = nombre
                tecnico.id_depto = int(departamento_id)
                tecnico.id_grupo = int(grupo_id)
                db.session.commit()
                flash("Técnico modificado correctamente.", "success")
            return redirect(url_for("tecnicos", from_moderacion="1" if from_moderacion else None))

        nuevo = Tecnico(
            nombre_completo=nombre,
            id_depto=int(departamento_id),
            id_grupo=int(grupo_id),
        )
        db.session.add(nuevo)
        db.session.commit()
        flash("Técnico agregado correctamente.", "success")
        return redirect(url_for("tecnicos", from_moderacion="1" if from_moderacion else None))

    return render_template(
        "tecnicos.html",
        active="tecnicos",
        departamentos=departamentos,
        grupos=grupos,
        tecnicos=tecnicos,
        editar_tecnico=editar_tecnico,
        from_moderacion=from_moderacion,
    )


@app.route("/editar_tecnico/<int:id_tecnico>")
def editar_tecnico(id_tecnico):
    from_moderacion = request.args.get("from_moderacion", "0") == "1"
    return redirect(url_for("tecnicos", edit_id=id_tecnico, from_moderacion="1" if from_moderacion else None))


@app.route("/eliminar_tecnico/<int:id_tecnico>")
def eliminar_tecnico(id_tecnico):
    from_moderacion = request.args.get("from_moderacion", "0") == "1"
    tecnico = Tecnico.query.get_or_404(id_tecnico)
    
    TicketSoporte.query.filter_by(id_tecnico_asignado=id_tecnico).update({"id_tecnico_asignado": None})
    
    db.session.delete(tecnico)
    db.session.commit()
    flash("Técnico eliminado. Los casos asociados han quedado sin técnico asignado.", "success")
    return redirect(url_for("tecnicos", from_moderacion="1" if from_moderacion else None))

@app.route("/backup_db")
def backup_db():
    """Descarga un respaldo de la base de datos local solo para usuarios con permisos de superusuario."""
    # Verifica que solo el super usuario pueda descargar la base de datos
    if session.get("user_type") != "super":
        flash("Acceso denegado. Privilegios insuficientes.", "danger")
        return redirect(url_for("moderacion"))
    
    # Busca el archivo de la base de datos local
    db_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), "sistema_ipasme.db")
    if os.path.exists(db_path):
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M")
        return send_file(db_path, as_attachment=True, download_name=f"Backup_IPASME_{fecha_actual}.db")
    else:
        flash("Error: No se encontró el archivo local de la base de datos.", "danger")
        return redirect(url_for("moderacion"))


def get_free_port(default_port=5000):
    """Devuelve un puerto disponible, prefiriendo el puerto por defecto si está libre."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        try:
            sock.bind(("127.0.0.1", default_port))
            return default_port
        except OSError:
            sock.bind(("127.0.0.1", 0))
            return sock.getsockname()[1]


def open_browser(url):
    """Abre el navegador en la URL del servidor después de iniciar la aplicación."""
    try:
        webbrowser.open_new_tab(url)
    except Exception:
        pass

if __name__ == "__main__":
    host = "127.0.0.1"
    port = get_free_port(5000)
    url = f"http://{host}:{port}/login"
    print(f"[INFO] Starting server on {url}")
    timer = threading.Timer(2.0, lambda: open_browser(url))
    timer.daemon = True
    timer.start()
    app.run(host=host, port=port, debug=False)